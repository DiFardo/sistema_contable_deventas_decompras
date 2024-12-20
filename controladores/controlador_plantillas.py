from flask import send_file, jsonify
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY, FORMAT_NUMBER_COMMA_SEPARATED1
from psycopg2.extras import DictCursor
from bd_conexion import obtener_conexion
from psycopg2 import sql, Error
from openpyxl import Workbook
from io import BytesIO
from flask import send_file
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from sqlalchemy import create_engine, text
from dotenv import load_dotenv
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import os
from reportlab.pdfgen import canvas
from io import BytesIO
from flask import send_file, jsonify
from reportlab.lib.units import inch

def generar_registro_venta_excel(mes, anio):
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor()

        consulta = """
        SELECT 
            ROW_NUMBER() OVER(ORDER BY v.serie_comprobante, v.numero_comprobante) AS correlativo,
            TO_CHAR(v.fecha, 'DD/MM/YYYY') AS fecha_emision,
            CASE WHEN v.tipo_comprobante = 'Boleta' THEN '03' ELSE '01' END AS tipo_comprobante,
            v.serie_comprobante,
            v.numero_comprobante,
            CASE 
                WHEN v.tipo_documento = 'DNI' THEN '1' 
                WHEN v.tipo_documento = 'Carnet de extranjería' THEN '4'
                WHEN v.tipo_documento = 'RUC' THEN '6'
                WHEN v.tipo_documento = 'Pasaporte' THEN '7'
                ELSE ''
            END AS tipo_documento,
            v.numero_documento,
            v.usuario,
            SUM(v.sub_sin_igv) AS base_imponible,
            SUM(v.igv) AS igv,
            SUM(v.total) AS total_comprobante
        FROM ventas_contables v
        WHERE EXTRACT(MONTH FROM v.fecha) = %s AND EXTRACT(YEAR FROM v.fecha) = %s
        GROUP BY v.serie_comprobante, v.numero_comprobante, v.tipo_documento, 
                 v.numero_documento, v.usuario, v.tipo_comprobante, v.fecha
        ORDER BY v.fecha, v.serie_comprobante, v.numero_comprobante;
        """
        cursor.execute(consulta, (mes, anio))

        resultados = cursor.fetchall()

        ruta_plantilla = 'plantillas/RegistroVentas.xlsx'
        workbook = load_workbook(ruta_plantilla)
        hoja = workbook.active
        
        periodo = f"{anio}-{str(mes).zfill(2)}"
        hoja.cell(row=3, column=2, value=periodo)

        borde = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        fuente_estandar = Font(name='Arial', size=10)

        fila_base = 12
        alto_fila_base = hoja.row_dimensions[fila_base].height

        fila_inicial = fila_base
        total_base_imponible = 0
        total_igv = 0
        total_comprobante = 0

        columnas_con_borde = list(range(1, 23))

        for fila, registro in enumerate(resultados, start=fila_inicial):
            hoja.row_dimensions[fila].height = alto_fila_base

            celdas = [
                (1, registro[0]),  # Correlativo
                (2, registro[1]),  # Fecha de emisión
                (4, registro[2]),  # Tipo de comprobante
                (5, registro[3]),  # Serie comprobante
                (6, registro[4]),  # Número comprobante
                (7, registro[5]),  # Tipo documento
                (8, registro[6]),  # Número documento
                (9, registro[7]),  # Usuario
                (11, registro[8]),  # Base imponible
                (15, registro[9]),  # IGV
                (17, registro[10])  # Total comprobante
            ]

            for col in columnas_con_borde:
                celda = hoja.cell(row=fila, column=col)
                celda.border = borde
                celda.font = fuente_estandar

            for col, valor in celdas:
                celda = hoja.cell(row=fila, column=col, value=valor)
                celda.font = fuente_estandar

                if col in (11, 15, 17):
                    celda.alignment = Alignment(horizontal='right', vertical='center')
                    celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                elif col == 2:
                    celda.number_format = FORMAT_DATE_DDMMYY
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                elif col in (1, 4, 5, 6, 7, 8):
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    celda.alignment = Alignment(horizontal='left', vertical='center')

                if col == 11:
                    total_base_imponible += valor
                elif col == 15:
                    total_igv += valor
                elif col == 17:
                    total_comprobante += valor

        # Escribir los totales en la fila de "Totales"
        fila_totales = fila_inicial + len(resultados)
        hoja.row_dimensions[fila_totales].height = alto_fila_base
        celda_totales = hoja.cell(row=fila_totales, column=9, value="TOTALES")
        celda_totales.border = borde
        celda_totales.alignment = Alignment(horizontal='center', vertical='center')
        celda_totales.font = Font(name='Arial', size=10, bold=True)

        total_base_imponible_celda = hoja.cell(row=fila_totales, column=11, value=total_base_imponible)
        total_base_imponible_celda.border = borde
        total_base_imponible_celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        total_base_imponible_celda.alignment = Alignment(horizontal='right', vertical='center')
        total_base_imponible_celda.font = fuente_estandar

        total_igv_celda = hoja.cell(row=fila_totales, column=15, value=total_igv)
        total_igv_celda.border = borde
        total_igv_celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        total_igv_celda.alignment = Alignment(horizontal='right', vertical='center')
        total_igv_celda.font = fuente_estandar

        total_comprobante_celda = hoja.cell(row=fila_totales, column=17, value=total_comprobante)
        total_comprobante_celda.border = borde
        total_comprobante_celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        total_comprobante_celda.alignment = Alignment(horizontal='right', vertical='center')
        total_comprobante_celda.font = fuente_estandar

        for col in columnas_con_borde:
            hoja.cell(row=fila_totales, column=col).border = borde

        # Guardar el archivo Excel en memoria
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        nombre_archivo = f'registro_ventas_{anio}_{mes}.xlsx'

        # Enviar el archivo como respuesta
        return send_file(
            output,
            download_name=nombre_archivo,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error al generar el registro de ventas: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close()
            conexion.close()

def generar_registro_compra_excel(mes, anio):
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor()

        consulta = """
            SELECT 
                ROW_NUMBER() OVER(ORDER BY c.serie_comprobante, c.numero_comprobante) AS correlativo,
                TO_CHAR(c.fecha, 'DD/MM/YYYY') AS fecha_emision,
                CASE WHEN c.tipo_comprobante = 'Factura' THEN '01' ELSE '03' END AS tipo_comprobante,
                c.serie_comprobante,
                c.numero_comprobante,
                CASE 
                    WHEN c.tipo_documento = 'DNI' THEN '1' 
                    WHEN c.tipo_documento = 'Carnet de extranjería' THEN '4'
                    WHEN c.tipo_documento = 'RUC' THEN '6'
                    WHEN c.tipo_documento = 'Pasaporte' THEN '7'
                    ELSE ''
                END AS tipo_documento,
                c.numero_documento,
                c.nombre_proveedor,
                SUM(c.sub_sin_igv) AS base_imponible,
                SUM(c.igv) AS igv,
                SUM(c.total) AS total_comprobante
            FROM compras_contables c
            WHERE EXTRACT(MONTH FROM c.fecha) = %s AND EXTRACT(YEAR FROM c.fecha) = %s
            GROUP BY c.serie_comprobante, c.numero_comprobante, c.tipo_documento, 
                     c.numero_documento, c.nombre_proveedor, c.tipo_comprobante, c.fecha
            ORDER BY c.fecha, c.serie_comprobante, c.numero_comprobante;
        """
        cursor.execute(consulta, (mes, anio))

        resultados = cursor.fetchall()

        ruta_plantilla = 'plantillas/RegistroCompras.xlsx'
        workbook = load_workbook(ruta_plantilla)
        hoja = workbook.active

        periodo = f"{anio}-{str(mes).zfill(2)}"
        hoja.cell(row=3, column=2, value=periodo)

        borde = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        fuente_estandar = Font(name='Arial', size=10)

        fila_base = 14
        alto_fila_base = hoja.row_dimensions[fila_base].height

        fila_inicial = fila_base
        total_base_imponible = 0
        total_igv = 0
        total_comprobante = 0

        columnas_con_borde = list(range(1, 29))

        for fila, registro in enumerate(resultados, start=fila_inicial):
            hoja.row_dimensions[fila].height = alto_fila_base

            celdas = [
                (1, registro[0]),  # Correlativo
                (2, registro[1]),  # Fecha de emisión
                (4, registro[2]),  # Tipo de comprobante
                (5, registro[3]),  # Serie comprobante
                (7, registro[4]),  # Número comprobante
                (8, registro[5]),  # Tipo documento
                (9, registro[6]),  # Número documento
                (10, registro[7]),  # Nombre proveedor
                (11, registro[8]),  # Base imponible
                (12, registro[9]),  # IGV
                (20, registro[10])  # Total comprobante
            ]

            for col in columnas_con_borde:
                celda = hoja.cell(row=fila, column=col)
                celda.border = borde
                celda.font = fuente_estandar

            for col, valor in celdas:
                celda = hoja.cell(row=fila, column=col, value=valor)
                celda.font = fuente_estandar

                if col in (11, 12, 20):
                    celda.alignment = Alignment(horizontal='right', vertical='center')
                    celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                elif col == 1:
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                elif col == 2:
                    celda.number_format = FORMAT_DATE_DDMMYY
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                elif col in (4, 5, 7, 8, 9):
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    celda.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                if col == 11:
                    total_base_imponible += valor
                elif col == 12:
                    total_igv += valor
                elif col == 20:
                    total_comprobante += valor

        # Escribir los totales en la fila de "Totales"
        fila_totales = fila_inicial + len(resultados)
        hoja.row_dimensions[fila_totales].height = alto_fila_base
        celda_totales = hoja.cell(row=fila_totales, column=10, value="TOTALES")
        celda_totales.border = borde
        celda_totales.alignment = Alignment(horizontal='center', vertical='center')
        celda_totales.font = Font(name='Arial', size=10, bold=True)

        total_base_imponible_celda = hoja.cell(row=fila_totales, column=11, value=total_base_imponible)
        total_base_imponible_celda.border = borde
        total_base_imponible_celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        total_base_imponible_celda.alignment = Alignment(horizontal='right', vertical='center')
        total_base_imponible_celda.font = fuente_estandar

        total_igv_celda = hoja.cell(row=fila_totales, column=12, value=total_igv)
        total_igv_celda.border = borde
        total_igv_celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        total_igv_celda.alignment = Alignment(horizontal='right', vertical='center')
        total_igv_celda.font = fuente_estandar

        total_comprobante_celda = hoja.cell(row=fila_totales, column=20, value=total_comprobante)
        total_comprobante_celda.border = borde
        total_comprobante_celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        total_comprobante_celda.alignment = Alignment(horizontal='right', vertical='center')
        total_comprobante_celda.font = fuente_estandar

        for col in columnas_con_borde:
            hoja.cell(row=fila_totales, column=col).border = borde

        # Guardar el archivo Excel en memoria
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        nombre_archivo = f'registro_compras_{anio}_{mes}.xlsx'

        # Enviar el archivo como respuesta
        return send_file(
            output,
            download_name=nombre_archivo,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error al generar el registro de compras: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close()
            conexion.close()

def generar_libro_diario_excel(fecha_inicio, fecha_fin=None):
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor()   
        if fecha_fin:
            consulta = """
                SELECT
                    DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                    TO_CHAR(ac.fecha, 'DD/MM/YYYY') AS fecha,
                    CASE
                        WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                        WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                        ELSE ''
                    END AS glosa,
                    CASE
                        WHEN m.tipo_movimiento = 'Compras' THEN 8
                        WHEN m.tipo_movimiento = 'Ventas' THEN 14
                        ELSE NULL
                    END AS codigo_del_libro,
                    DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo_documento,
                    ac.numero_documento AS numero_documento_sustentatorio,
                    ac.codigo_cuenta,
                    ac.denominacion,
                    ac.debe,
                    ac.haber
                FROM asientos_contables ac
                JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
                WHERE ac.fecha::date BETWEEN %s AND %s
                ORDER BY ac.fecha ASC, numero_correlativo, ac.id;
            """
            cursor.execute(consulta, (fecha_inicio, fecha_fin))
        else:
            consulta = """
                SELECT
                    DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                    TO_CHAR(ac.fecha, 'DD/MM/YYYY') AS fecha,
                    CASE
                        WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                        WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                        ELSE ''
                    END AS glosa,
                    CASE
                        WHEN m.tipo_movimiento = 'Compras' THEN 8
                        WHEN m.tipo_movimiento = 'Ventas' THEN 14
                        ELSE NULL
                    END AS codigo_del_libro,
                    DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo_documento,
                    ac.numero_documento AS numero_documento_sustentatorio,
                    ac.codigo_cuenta,
                    ac.denominacion,
                    ac.debe,
                    ac.haber
                FROM asientos_contables ac
                JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
                WHERE ac.fecha::date = %s
                ORDER BY ac.fecha ASC, numero_correlativo, ac.id;
            """
            cursor.execute(consulta, (fecha_inicio,))
        
        resultados = cursor.fetchall()

        # Cargar la plantilla de Excel
        ruta_plantilla = 'plantillas/LibroDiario.xlsx'
        workbook = load_workbook(ruta_plantilla)
        hoja = workbook.active

        # Estilos y formatos
        borde = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        fuente_estandar = Font(name='Arial', size=10)

        # Actualizar la fecha en el Excel
        if fecha_fin:
            rango_fechas = f"{fecha_inicio} al {fecha_fin}"
        else:
            rango_fechas = fecha_inicio
        hoja.cell(row=3, column=2, value=rango_fechas)

        fila_base = 11
        alto_fila_base = hoja.row_dimensions[fila_base].height

        fila_inicial = fila_base
        total_debe = 0
        total_haber = 0

        columnas_con_borde = list(range(1, 11))

        for fila, registro in enumerate(resultados, start=fila_inicial):
            hoja.row_dimensions[fila].height = alto_fila_base

            celdas = [
                (1, registro[0]),  # Número correlativo
                (2, registro[1]),  # Fecha
                (3, registro[2]),  # Glosa
                (4, registro[3]),  # Código del libro
                (5, registro[4]),  # Número correlativo documento
                (6, registro[5]),  # Número documento sustentatorio
                (7, registro[6]),  # Código cuenta
                (8, registro[7]),  # Denominación
                (9, registro[8]),  # Debe
                (10, registro[9])  # Haber
            ]

            for col in columnas_con_borde:
                celda = hoja.cell(row=fila, column=col)
                celda.border = borde
                celda.font = fuente_estandar

            for col, valor in celdas:
                celda = hoja.cell(row=fila, column=col, value=valor)
                celda.font = fuente_estandar

                if col in (9, 10):  # Columnas "Debe" y "Haber"
                    celda.alignment = Alignment(horizontal='right', vertical='center')
                    celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                elif col in (1, 4, 5, 6):  # Columnas centradas
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                elif col == 2:  # Fecha
                    celda.number_format = FORMAT_DATE_DDMMYY
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                else:  # Otras columnas
                    celda.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                # Sumar totales
                if col == 9:
                    total_debe += valor or 0
                elif col == 10:
                    total_haber += valor or 0

        # Escribir los totales
        fila_totales = fila_inicial + len(resultados)
        hoja.row_dimensions[fila_totales].height = alto_fila_base
        celda_totales = hoja.cell(row=fila_totales, column=8, value="TOTALES")
        celda_totales.border = borde
        celda_totales.alignment = Alignment(horizontal='center', vertical='center')
        celda_totales.font = Font(name='Arial', size=10, bold=True)

        total_debe_celda = hoja.cell(row=fila_totales, column=9, value=total_debe)
        total_debe_celda.border = borde
        total_debe_celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        total_debe_celda.alignment = Alignment(horizontal='right', vertical='center')
        total_debe_celda.font = fuente_estandar

        total_haber_celda = hoja.cell(row=fila_totales, column=10, value=total_haber)
        total_haber_celda.border = borde
        total_haber_celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        total_haber_celda.alignment = Alignment(horizontal='right', vertical='center')
        total_haber_celda.font = fuente_estandar

        for col in columnas_con_borde:
            hoja.cell(row=fila_totales, column=col).border = borde

        # Guardar el archivo en memoria
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Generar el nombre del archivo
        nombre_archivo = f'libro_diario_{fecha_inicio}'
        if fecha_fin:
            nombre_archivo += f'_al_{fecha_fin}'
        nombre_archivo += '.xlsx'

        # Enviar el archivo como descarga
        return send_file(
            output,
            download_name=nombre_archivo,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error al generar el libro diario: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close()
            conexion.close()

            

# Función para agregar marca de agua con una imagen
def add_watermark_image(canvas, doc):
    # Cargar la imagen (asegúrate de tener la imagen en tu servidor o proyecto)
    watermark_image = "static/img/illustrations/logoEquipo.png"  # Cambia esta ruta a tu imagen
    
    canvas.saveState()
    
    # Establecer la posición y tamaño de la imagen de la marca de agua
    width, height = landscape(A4)  # Obtener el tamaño de la página
    image_width = width * 0.6  # Ajusta el ancho de la marca de agua (60% de la página)
    image_height = height * 0.6  # Ajusta el alto de la marca de agua (60% de la página)
    
    # Posición de la imagen (centrada)
    x_position = (width - image_width) / 2
    y_position = (height - image_height) / 2
    
    # Dibujar la imagen con opacidad
    canvas.setFillColor(colors.grey, alpha=0.2)  # Opacidad al 10%
    canvas.drawImage(watermark_image, x_position, y_position, width=image_width, height=image_height, mask='auto')
    
    canvas.restoreState()

def generar_libro_diario_pdf_horizontal(fecha_inicio, fecha_fin=None):
    try:
        if fecha_fin:
            movimientos_agrupados, total_debe, total_haber = obtener_libro_diario_por_fecha(fecha_inicio, fecha_fin)
        else:
            movimientos_agrupados, total_debe, total_haber = obtener_libro_diario_por_fecha(fecha_inicio)

        if not movimientos_agrupados:
            return jsonify({'error': 'No se encontraron datos para las fechas seleccionadas.'}), 404

        # Crear buffer para el PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),  # Configuración horizontal
            rightMargin=10, leftMargin=10, topMargin=10, bottomMargin=10  # Márgenes reducidos
        )

        # Crear estilos
        styles = getSampleStyleSheet()
        style_title = styles["Title"]
        style_normal = styles["Normal"]
        style_normal.fontSize = 9
        style_normal.leading = 11

        # Datos de la cabecera
        ruc = "20612188930"  # RUC fijo
        razon_social = "DECO ELERA S.A.C."  # Razón Social

        # Determinar el rango de fechas para el título
        fecha_texto = f"{fecha_inicio}" if not fecha_fin else f"{fecha_inicio} al {fecha_fin}"

        # Agregar título y cabecera
        elementos = []

        # Título del documento
        titulo = Paragraph(f"<b>Libro Diario - Detalle de las operaciones</b>", style_title)
        elementos.append(titulo)
        elementos.append(Spacer(1, 12))  # Espacio después del título

        # Cabecera con Periodo, RUC, Razón Social
        cabecera = [
            f"Periodo: {fecha_texto}",  # Periodo con fecha o rango de fechas
            f"RUC: {ruc}",  # RUC
            f"Razón Social: {razon_social}"  # Razón Social
        ]
        
        # Modificar el tamaño de fuente para las cabeceras
        style_header = styles["Normal"]
        style_header.fontSize = 11
        style_header.leading = 14

        for linea in cabecera:
            cabecera_parrafo = Paragraph(f"<b>{linea}</b>", style_header)
            elementos.append(cabecera_parrafo)
            elementos.append(Spacer(1, 6))  # Espacio entre las líneas de la cabecera

        # Espaciado antes de la tabla
        elementos.append(Spacer(1, 12))

        # Crear la tabla con los encabezados
        encabezados = [
            "N°", "Fecha", "Descripción", "Código",
            "N° Documento", "Código Cuenta", "Denominación", "Debe", "Haber"
        ]
        tabla_datos = [encabezados]

        # Procesar los datos agrupados
        for movimiento in movimientos_agrupados:
            # Encabezado del movimiento
            primera_fila = [
                movimiento["numero_correlativo"],  # N° Correlativo
                movimiento["fecha"],  # Fecha
                movimiento["glosa"],  # Descripción de la operación
                movimiento["codigo_del_libro"] or "-",  # Código del libro
                movimiento["numero_documento_sustentatorio"] or "-",  # N° Documento Sustentatorio
                "",  # Código Cuenta vacío (para que no deje espacio innecesario)
                "",  # Denominación vacía
                "",  # Debe
                ""   # Haber
            ]
            tabla_datos.append(primera_fila)

            # Cuentas asociadas al movimiento
            for idx, cuenta in enumerate(movimiento["cuentas"]):
                cuenta_fila = [
                    "",  # N° Correlativo vacío
                    "",  # Fecha vacía
                    "",  # Descripción vacía
                    "",  # Código del libro vacío
                    "",  # N° Documento vacío
                    cuenta["codigo_cuenta"] or "-",  # Código Cuenta
                    Paragraph(cuenta["denominacion"] or "-", style_normal),  # Denominación con ajuste
                    f"{cuenta['debe']:,.2f}" if cuenta["debe"] else "-",  # Debe
                    f"{cuenta['haber']:,.2f}" if cuenta["haber"] else "-"  # Haber
                ]
                tabla_datos.append(cuenta_fila)

            # Separador vacío entre movimientos
            tabla_datos.append([""] * len(encabezados))  # Solo se añade el separador vacío entre movimientos

        # Agregar fila de totales
        tabla_datos.append(["", "", "", "", "", "", "Totales", f"{total_debe:,.2f}", f"{total_haber:,.2f}"])

        # Ajustar tamaños de columna para ocupar más espacio
        tabla = Table(
            tabla_datos,
            colWidths=[50, 70, 150, 50, 100, 60, 220, 70, 70]  # Columnas ajustadas
        )
        tabla.setStyle(TableStyle([ 
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Fondo gris para encabezados
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Texto blanco para encabezados
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alineación general centrada
            ('ALIGN', (5, 1), (5, -1), 'LEFT'),  # Código Cuenta alineado a la izquierda
            ('ALIGN', (6, 1), (6, -1), 'LEFT'),  # Denominación alineado a la izquierda
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fuente en negrita para encabezados
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Fuente normal para contenido
            ('FONTSIZE', (0, 1), (-1, -1), 9),  # Fuente pequeña para contenido
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),  # Espaciado inferior en encabezados
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Líneas de cuadrícula
            ('VALIGN', (6, 1), (6, -1), 'MIDDLE'),  # Alinear verticalmente Denominación al centro
            ('SPAN', (0, -1), (5, -1)),  # Fusionar celdas para Totales
            ('ALIGN', (-2, 1), (-1, -1), 'RIGHT'),  # Alinear Debe y Haber a la derecha
        ]))
        elementos.append(tabla)

        # Generar el PDF con la marca de agua en cada página
        doc.build(elementos, onFirstPage=add_watermark_image, onLaterPages=add_watermark_image)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"Libro_Diario_{fecha_inicio}_to_{fecha_fin or fecha_inicio}.pdf",
            mimetype="application/pdf"
        )
    except Exception as e:
        print(f"Error al generar el PDF: {e}")
        return jsonify({"error": str(e)}), 500



def generar_libro_caja_pdf_horizontal(mes, año):
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor(cursor_factory=DictCursor)

        consulta = """
            SELECT
                DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                TO_CHAR(ac.fecha, 'DD/MM/YYYY') AS fecha_operacion,
                CASE
                    WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                    WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                    ELSE 'Descripción no especificada'
                END AS descripcion_operacion,
                ac.codigo_cuenta AS codigo_cuenta_asociada,
                ac.denominacion AS denominacion_cuenta_asociada,
                COALESCE(ac.debe, 0) AS saldo_deudor,
                COALESCE(ac.haber, 0) AS saldo_acreedor
            FROM asientos_contables ac
            JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
            WHERE EXTRACT(MONTH FROM ac.fecha) = %s AND EXTRACT(YEAR FROM ac.fecha) = %s
            ORDER BY fecha_operacion, numero_correlativo, ac.id;
        """
        cursor.execute(consulta, (mes, año))
        resultados = cursor.fetchall()

        if not resultados:
            return jsonify({'error': 'No se encontraron datos para el periodo seleccionado.'}), 404

        # Procesar datos y ajustar capitalización en denominación
        procesados = []
        for registro in resultados:
            denominación = registro["denominacion_cuenta_asociada"]
            if denominación:
                denominación = denominación.capitalize()
            registro_procesado = (
                registro["numero_correlativo"],
                registro["fecha_operacion"],
                registro["descripcion_operacion"],
                registro["codigo_cuenta_asociada"],
                denominación,
                registro["saldo_deudor"],
                registro["saldo_acreedor"]
            )
            procesados.append(registro_procesado)

        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=10,
            leftMargin=10,
            topMargin=20,
            bottomMargin=20
        )

        # Estilos para el documento
        styles = getSampleStyleSheet()
        style_title = styles["Title"]
        style_normal = styles["Normal"]
        style_normal.fontSize = 9
        style_normal.leading = 11

        # Título del documento (con el texto adecuado para el Libro Caja)
        titulo = Paragraph(f"<b>Libro Caja y Bancos - Detalle de los movimientos del efectivo</b>", style_title)
        
        # Cabecera con el periodo, RUC y Razón Social
        ruc = "20612188930"  # RUC fijo
        razon_social = "DECO ELERA S.A.C."  # Razón Social
        fecha_texto = f"{mes}/{año}"

        cabecera = [
            f"Periodo: {fecha_texto}",  # Periodo con fecha o rango de fechas
            f"RUC: {ruc}",  # RUC
            f"Razón Social: {razon_social}"  # Razón Social
        ]
        
        # Agregar título y cabecera al documento
        elementos = []
        elementos.append(titulo)
        elementos.append(Spacer(1, 12))  # Espacio después del título

        # Agregar la cabecera con el periodo, RUC y Razón Social
        style_header = styles["Normal"]
        style_header.fontSize = 11
        style_header.leading = 14

        for linea in cabecera:
            cabecera_parrafo = Paragraph(f"<b>{linea}</b>", style_header)
            elementos.append(cabecera_parrafo)
            elementos.append(Spacer(1, 6))  # Espacio entre las líneas de la cabecera

        # Espaciado antes de la tabla
        elementos.append(Spacer(1, 12))

        # Crear la tabla con los encabezados
        encabezados = ["N°", "Fecha", "Descripción", "Código", "Denominación", "Deudor", "Acreedor"]
        tabla_datos = [encabezados]

        for registro in procesados:
            fila = [
                str(registro[0]),
                registro[1],
                registro[2],
                str(registro[3] or "-"),
                registro[4],
                f"{registro[5]:,.2f}" if registro[5] else "-",
                f"{registro[6]:,.2f}" if registro[6] else "-"
            ]
            tabla_datos.append(fila)

        total_deudor = sum([registro[5] or 0 for registro in procesados])
        total_acreedor = sum([registro[6] or 0 for registro in procesados])
        tabla_datos.append(["", "", "", "", "Totales", f"{total_deudor:,.2f}", f"{total_acreedor:,.2f}"])

        # Estilizar tabla
        tabla = Table(
            tabla_datos,
            colWidths=[30, 50, 120, 60, 300, 70, 70]
        )
        tabla.setStyle(TableStyle([ 
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),  # Descripción a la izquierda
            ('ALIGN', (3, 1), (3, -1), 'LEFT'),  # Código a la izquierda
            ('ALIGN', (4, 1), (4, -1), 'LEFT'),  # Denominación a la izquierda
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ALIGN', (-2, 1), (-1, -1), 'RIGHT'),  # Deudor y Acreedor a la derecha
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))

        elementos.append(tabla)

        # Generar el PDF con la marca de agua en cada página
        doc.build(elementos, onFirstPage=add_watermark_image, onLaterPages=add_watermark_image)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"Libro_Caja_{mes}_{año}.pdf",
            mimetype="application/pdf"
        )

    except Exception as e:
        print(f"Error al generar el PDF: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conexion:
            cursor.close()
            conexion.close()


def generar_registro_ventas_pdf(mes, año):
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor(cursor_factory=DictCursor)

        consulta = """
            SELECT 
                ROW_NUMBER() OVER(ORDER BY v.serie_comprobante, v.numero_comprobante) AS correlativo,
                TO_CHAR(v.fecha, 'DD/MM/YYYY') AS fecha_emision,
                CASE WHEN v.tipo_comprobante = 'Boleta' THEN '03' ELSE '01' END AS tipo_comprobante,
                v.serie_comprobante,
                v.numero_comprobante,
                CASE 
                    WHEN v.tipo_documento = 'DNI' THEN '1' 
                    WHEN v.tipo_documento = 'Carnet de extranjería' THEN '4'
                    WHEN v.tipo_documento = 'RUC' THEN '6'
                    WHEN v.tipo_documento = 'Pasaporte' THEN '7'
                    ELSE ''
                END AS tipo_documento,
                v.numero_documento,
                v.usuario,
                SUM(v.sub_sin_igv) AS base_imponible,
                SUM(v.igv) AS igv,
                SUM(v.total) AS total_comprobante
            FROM ventas_contables v
            WHERE EXTRACT(MONTH FROM v.fecha) = %s AND EXTRACT(YEAR FROM v.fecha) = %s
            GROUP BY v.serie_comprobante, v.numero_comprobante, v.tipo_documento, 
                     v.numero_documento, v.usuario, v.tipo_comprobante, v.fecha
            ORDER BY v.fecha, v.serie_comprobante, v.numero_comprobante;
        """
        cursor.execute(consulta, (mes, año))
        resultados = cursor.fetchall()

        if not resultados:
            return jsonify({'error': 'No se encontraron datos para el período seleccionado.'}), 404

        # Procesar datos
        procesados = []
        for registro in resultados:
            razon_social = registro["usuario"]  # Dejar tal cual viene de la base de datos
            registro_procesado = (
                registro["correlativo"],
                registro["fecha_emision"],
                registro["tipo_comprobante"],
                registro["serie_comprobante"],
                registro["numero_comprobante"],
                registro["tipo_documento"],
                registro["numero_documento"],
                razon_social,  # No se altera la capitalización
                registro["base_imponible"],
                registro["igv"],
                registro["total_comprobante"]
            )
            procesados.append(registro_procesado)

        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=10,
            leftMargin=10,
            topMargin=20,
            bottomMargin=20
        )

        styles = getSampleStyleSheet()
        style_title = styles["Title"]
        style_normal = styles["Normal"]
        style_normal.fontSize = 9
        style_normal.leading = 11

        # Elementos para el PDF
        elementos = []

        # Título del documento
        titulo = Paragraph(f"<b>Registro de Ventas e Ingresos</b>", style_title)
        elementos.append(titulo)
        elementos.append(Spacer(1, 12))  # Espacio después del título

        # Cabecera con el Periodo, RUC, Razón Social
        ruc = "20612188930"  # RUC fijo
        razon_social = "DECO ELERA S.A.C."  # Razón Social
        fecha_texto = f"{mes}/{año}"

        cabecera = [
            f"Periodo: {fecha_texto}",  # Periodo con fecha o rango de fechas
            f"RUC: {ruc}",  # RUC
            f"Razón Social: {razon_social}"  # Razón Social
        ]
        
        # Agregar título y cabecera al documento
        style_header = styles["Normal"]
        style_header.fontSize = 11
        style_header.leading = 14

        for linea in cabecera:
            cabecera_parrafo = Paragraph(f"<b>{linea}</b>", style_header)
            elementos.append(cabecera_parrafo)
            elementos.append(Spacer(1, 6))  # Espacio entre las líneas de la cabecera

        # Espaciado antes de la tabla
        elementos.append(Spacer(1, 12))

        # Construir la tabla
        encabezados = [
            "N°", "Fecha Emisión", "Tipo Comp.", "Serie", "N° Comprobante",
            "Tipo Doc.", "N° Doc.", "Razón Social", "Base Imponible", "IGV", "Total"
        ]
        tabla_datos = [encabezados]

        for registro in procesados:
            fila = [
                str(registro[0]),
                registro[1],
                registro[2],
                registro[3],
                registro[4],
                registro[5],
                registro[6],
                registro[7],
                f"{registro[8]:,.2f}" if registro[8] else "-",
                f"{registro[9]:,.2f}" if registro[9] else "-",
                f"{registro[10]:,.2f}" if registro[10] else "-"
            ]
            tabla_datos.append(fila)

        # Totales
        total_base_imponible = sum([registro[8] or 0 for registro in procesados])
        total_igv = sum([registro[9] or 0 for registro in procesados])
        total_total_comprobante = sum([registro[10] or 0 for registro in procesados])
        tabla_datos.append(
            ["", "", "", "", "", "", "", "Totales", 
             f"{total_base_imponible:,.2f}", f"{total_igv:,.2f}", f"{total_total_comprobante:,.2f}"]
        )

        # Estilizar tabla
        tabla = Table(
            tabla_datos,
            colWidths=[30, 60, 60, 50, 70, 50, 70, 120, 80, 60, 60]
        )
        tabla.setStyle(TableStyle([ 
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (7, 1), (7, -1), 'LEFT'),  # Razón Social a la izquierda
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ALIGN', (-3, 1), (-1, -1), 'RIGHT'),  # Base Imponible, IGV y Total a la derecha
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))

        elementos.append(tabla)

        # Generar el PDF con la marca de agua en cada página
        doc.build(elementos, onFirstPage=add_watermark_image, onLaterPages=add_watermark_image)

        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"Registro_Ventas_{mes}_{año}.pdf",
            mimetype="application/pdf"
        )

    except Exception as e:
        print(f"Error al generar el PDF: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conexion:
            cursor.close()
            conexion.close()


def obtener_registro_ventas(mes, año):
    conexion = obtener_conexion()
    registros = []
    total_base_imponible = 0
    total_igv = 0
    total_total_comprobante = 0
    with conexion.cursor(cursor_factory=DictCursor) as cursor:
        cursor.execute("""
            SELECT 
                ROW_NUMBER() OVER(ORDER BY v.serie_comprobante, v.numero_comprobante) AS correlativo,
                TO_CHAR(v.fecha, 'DD/MM/YYYY') AS fecha_emision,
                CASE WHEN v.tipo_comprobante = 'Boleta' THEN '03' ELSE '01' END AS tipo_comprobante,
                v.serie_comprobante,
                v.numero_comprobante,
                CASE 
                    WHEN v.tipo_documento = 'DNI' THEN '1' 
                    WHEN v.tipo_documento = 'Carnet de extranjería' THEN '4'
                    WHEN v.tipo_documento = 'RUC' THEN '6'
                    WHEN v.tipo_documento = 'Pasaporte' THEN '7'
                    ELSE ''
                END AS tipo_documento,
                v.numero_documento,
                v.usuario,
                SUM(v.sub_sin_igv) AS base_imponible,
                SUM(v.igv) AS igv,
                SUM(v.total) AS total_comprobante
            FROM ventas_contables v
            WHERE EXTRACT(MONTH FROM v.fecha) = %s AND EXTRACT(YEAR FROM v.fecha) = %s
            GROUP BY v.serie_comprobante, v.numero_comprobante, v.tipo_documento, 
                     v.numero_documento, v.usuario, v.tipo_comprobante, v.fecha
            ORDER BY v.fecha, v.serie_comprobante, v.numero_comprobante;
        """, (mes, año))
        registros = cursor.fetchall()
        for registro in registros:
            total_base_imponible += registro['base_imponible']
            total_igv += registro['igv']
            total_total_comprobante += registro['total_comprobante']
    conexion.close()
    return registros, total_base_imponible, total_igv, total_total_comprobante

# Esta función realiza la consulta a la base de datos con el filtro de mes y año
def obtener_registro_ventas_por_fecha(mes, anio):
    conexion = obtener_conexion()
    registros = []
    total_base_imponible = 0
    total_igv = 0
    total_total_comprobante = 0
    with conexion.cursor(cursor_factory=DictCursor) as cursor:
        try:
            cursor.execute("""
                SELECT 
                    ROW_NUMBER() OVER(ORDER BY v.serie_comprobante, v.numero_comprobante) AS correlativo,
                    TO_CHAR(v.fecha, 'DD/MM/YYYY') AS fecha_emision,
                    CASE WHEN v.tipo_comprobante = 'Boleta' THEN '03' ELSE '01' END AS tipo_comprobante,
                    v.serie_comprobante,
                    v.numero_comprobante,
                    CASE 
                        WHEN v.tipo_documento = 'DNI' THEN '1' 
                        WHEN v.tipo_documento = 'Carnet de extranjería' THEN '4'
                        WHEN v.tipo_documento = 'RUC' THEN '6'
                        WHEN v.tipo_documento = 'Pasaporte' THEN '7'
                        ELSE ''
                    END AS tipo_documento,
                    v.numero_documento,
                    v.usuario,
                    SUM(v.sub_sin_igv) AS base_imponible,
                    SUM(v.igv) AS igv,
                    SUM(v.total) AS total_comprobante
                FROM ventas_contables v
                WHERE EXTRACT(MONTH FROM v.fecha) = %s AND EXTRACT(YEAR FROM v.fecha) = %s
                GROUP BY v.serie_comprobante, v.numero_comprobante, v.tipo_documento, 
                         v.numero_documento, v.usuario, v.tipo_comprobante, v.fecha
                ORDER BY v.fecha, v.serie_comprobante, v.numero_comprobante;
            """, (mes, anio))
            registros = cursor.fetchall()
            for registro in registros:
                total_base_imponible += registro['base_imponible']
                total_igv += registro['igv']
                total_total_comprobante += registro['total_comprobante']
            print(f"Registros obtenidos: {registros}")
            print(f"Total Base Imponible: {total_base_imponible}, Total IGV: {total_igv}, Total Comprobante: {total_total_comprobante}")
        except Exception as e:
            print(f"Error al obtener registros de ventas: {e}")
    conexion.close()
    return registros, total_base_imponible, total_igv, total_total_comprobante


def obtener_registro_compras(mes, año):
    conexion = obtener_conexion()
    registros = []
    total_base_imponible = 0
    total_igv = 0
    total_total_comprobante = 0

    with conexion.cursor(cursor_factory=DictCursor) as cursor:
        cursor.execute("""
            SELECT 
                ROW_NUMBER() OVER(ORDER BY c.serie_comprobante, c.numero_comprobante) AS correlativo,
                TO_CHAR(c.fecha, 'DD/MM/YYYY') AS fecha_emision,
                CASE WHEN c.tipo_comprobante = 'Factura' THEN '01' ELSE '03' END AS tipo_comprobante,
                c.serie_comprobante,
                c.numero_comprobante,
                CASE 
                    WHEN c.tipo_documento = 'DNI' THEN '1' 
                    WHEN c.tipo_documento = 'Carnet de extranjería' THEN '4'
                    WHEN c.tipo_documento = 'RUC' THEN '6'
                    WHEN c.tipo_documento = 'Pasaporte' THEN '7'
                    ELSE ''
                END AS tipo_documento,
                c.numero_documento,
                c.nombre_proveedor,
                SUM(c.sub_sin_igv) AS base_imponible,
                SUM(c.igv) AS igv,
                SUM(c.total) AS total_comprobante
            FROM compras_contables c
            WHERE EXTRACT(MONTH FROM c.fecha) = %s AND EXTRACT(YEAR FROM c.fecha) = %s
            GROUP BY c.serie_comprobante, c.numero_comprobante, c.tipo_documento, 
                     c.numero_documento, c.nombre_proveedor, c.tipo_comprobante, c.fecha
            ORDER BY c.fecha, c.serie_comprobante, c.numero_comprobante;
        """, (mes, año))

        registros = cursor.fetchall()

        for registro in registros:
            total_base_imponible += registro['base_imponible']
            total_igv += registro['igv']
            total_total_comprobante += registro['total_comprobante']

    conexion.close()

    return registros, total_base_imponible, total_igv, total_total_comprobante


def generar_registro_compras_pdf(mes, año):
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor(cursor_factory=DictCursor)

        consulta = """
            SELECT 
                ROW_NUMBER() OVER(ORDER BY c.serie_comprobante, c.numero_comprobante) AS correlativo,
                TO_CHAR(c.fecha, 'DD/MM/YYYY') AS fecha_emision,
                CASE WHEN c.tipo_comprobante = 'Factura' THEN '01' ELSE '03' END AS tipo_comprobante,
                c.serie_comprobante,
                c.numero_comprobante,
                CASE 
                    WHEN c.tipo_documento = 'DNI' THEN '1' 
                    WHEN c.tipo_documento = 'Carnet de extranjería' THEN '4'
                    WHEN c.tipo_documento = 'RUC' THEN '6'
                    WHEN c.tipo_documento = 'Pasaporte' THEN '7'
                    ELSE ''
                END AS tipo_documento,
                c.numero_documento,
                c.nombre_proveedor,
                SUM(c.sub_sin_igv) AS base_imponible,
                SUM(c.igv) AS igv,
                SUM(c.total) AS total_comprobante
            FROM compras_contables c
            WHERE EXTRACT(MONTH FROM c.fecha) = %s AND EXTRACT(YEAR FROM c.fecha) = %s
            GROUP BY c.serie_comprobante, c.numero_comprobante, c.tipo_documento, 
                     c.numero_documento, c.nombre_proveedor, c.tipo_comprobante, c.fecha
            ORDER BY c.fecha, c.serie_comprobante, c.numero_comprobante;
        """
        cursor.execute(consulta, (mes, año))
        resultados = cursor.fetchall()

        if not resultados:
            return jsonify({'error': 'No se encontraron datos para el período seleccionado.'}), 404

        # Procesar datos
        procesados = []
        for registro in resultados:
            procesados.append(
                (
                    registro["correlativo"],
                    registro["fecha_emision"],
                    registro["tipo_comprobante"],
                    registro["serie_comprobante"],
                    registro["numero_comprobante"],
                    registro["tipo_documento"],
                    registro["numero_documento"],
                    Paragraph(registro["nombre_proveedor"], getSampleStyleSheet()['BodyText']),
                    registro["base_imponible"],
                    registro["igv"],
                    registro["total_comprobante"],
                )
            )

        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=10,
            leftMargin=10,
            topMargin=20,
            bottomMargin=20
        )

        styles = getSampleStyleSheet()
        style_title = styles["Title"]
        
        # Título del documento
        titulo = Paragraph(f"<b>Registro de Compras</b>", style_title)
        
        # Cabecera con el Periodo, RUC, Razón Social en negrita
        ruc = "20612188930"  # RUC fijo
        razon_social = "DECO ELERA S.A.C."  # Razón Social
        fecha_texto = f"{mes}/{año}"

        # Usar <b> para negrita en los textos de la cabecera
        cabecera = [
            f"<b>Periodo:{fecha_texto}</b> ",
            f"<b>RUC:{ruc}</b> ",
            f"<b>Razón Social: {razon_social}</b>"
        ]

        # Añadir título y cabecera a los elementos
        elementos = [titulo, Spacer(1, 12)]  # Espacio después del título

        for linea in cabecera:
            elementos.append(Paragraph(linea, styles['Normal']))
        
        elementos.append(Spacer(1, 12))  # Espacio después de la cabecera

        # Construir la tabla
        encabezados = [
            "N°", "Fecha", "Tipo Comprob.", "Serie", "N° Comprob.",
            "Tipo Doc.", "N° Doc.", "Proveedor", "Base Impon.", "IGV", "Total"
        ]
        tabla_datos = [encabezados]

        for registro in procesados:
            fila = [
                str(registro[0]),
                registro[1],
                registro[2],
                registro[3],
                registro[4],
                registro[5],
                registro[6],
                registro[7],
                f"{registro[8]:,.2f}" if registro[8] else "-",
                f"{registro[9]:,.2f}" if registro[9] else "-",
                f"{registro[10]:,.2f}" if registro[10] else "-",
            ]
            tabla_datos.append(fila)

        # Totales
        total_base_imponible = sum([registro[8] or 0 for registro in procesados])
        total_igv = sum([registro[9] or 0 for registro in procesados])
        total_total_comprobante = sum([registro[10] or 0 for registro in procesados])
        tabla_datos.append(
            ["", "", "", "", "", "", "", "Totales",
             f"{total_base_imponible:,.2f}", f"{total_igv:,.2f}", f"{total_total_comprobante:,.2f}"]
        )

        # Configurar tabla
        tabla = Table(
            tabla_datos,
            colWidths=[30, 60, 60, 50, 70, 50, 70, 180, 60, 60, 60]  # Aumentado el ancho para "Proveedor"
        )
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (7, 1), (7, -1), 'LEFT'),  # Proveedor alineado a la izquierda
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))

        # Añadir tabla a los elementos
        elementos.append(tabla)

        # Generar el PDF con la marca de agua en cada página
        doc.build(elementos, onFirstPage=add_watermark_image, onLaterPages=add_watermark_image)

        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"Registro_Compras_{mes}_{año}.pdf",
            mimetype="application/pdf"
        )

    except Exception as e:
        print(f"Error al generar el PDF: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conexion:
            cursor.close()
            conexion.close()


def obtener_registro_compras_por_fecha(mes, anio):
    conexion = obtener_conexion()
    registros = []
    total_base_imponible = 0
    total_igv = 0
    total_total_comprobante = 0

    with conexion.cursor(cursor_factory=DictCursor) as cursor:
        try:
            cursor.execute("""
                SELECT 
                    ROW_NUMBER() OVER(ORDER BY c.serie_comprobante, c.numero_comprobante) AS correlativo,
                    TO_CHAR(c.fecha, 'DD/MM/YYYY') AS fecha_emision,
                    CASE WHEN c.tipo_comprobante = 'Factura' THEN '01' ELSE '03' END AS tipo_comprobante,
                    c.serie_comprobante,
                    c.numero_comprobante,
                    CASE 
                        WHEN c.tipo_documento = 'DNI' THEN '1' 
                        WHEN c.tipo_documento = 'Carnet de extranjería' THEN '4'
                        WHEN c.tipo_documento = 'RUC' THEN '6'
                        WHEN c.tipo_documento = 'Pasaporte' THEN '7'
                        ELSE ''
                    END AS tipo_documento,
                    c.numero_documento,
                    c.nombre_proveedor,
                    SUM(c.sub_sin_igv) AS base_imponible,
                    SUM(c.igv) AS igv,
                    SUM(c.total) AS total_comprobante
                FROM compras_contables c
                WHERE EXTRACT(MONTH FROM c.fecha) = %s AND EXTRACT(YEAR FROM c.fecha) = %s
                GROUP BY c.serie_comprobante, c.numero_comprobante, c.tipo_documento, 
                         c.numero_documento, c.nombre_proveedor, c.tipo_comprobante, c.fecha
                ORDER BY c.fecha, c.serie_comprobante, c.numero_comprobante;
            """, (mes, anio))

            registros = cursor.fetchall()

            for registro in registros:
                total_base_imponible += registro['base_imponible']
                total_igv += registro['igv']
                total_total_comprobante += registro['total_comprobante']

            # Log output for debugging
            print(f"Registros obtenidos: {registros}")
            print(f"Total Base Imponible: {total_base_imponible}, Total IGV: {total_igv}, Total Comprobante: {total_total_comprobante}")

        except Exception as e:
            print(f"Error al obtener registros de compras: {e}")

    conexion.close()

    return registros, total_base_imponible, total_igv, total_total_comprobante


def obtener_libro_diario(fecha):
    conexion = obtener_conexion()
    movimientos = []
    total_debe = 0
    total_haber = 0

    with conexion.cursor(cursor_factory=DictCursor) as cursor:
        cursor.execute("""
            SELECT
                DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                ac.fecha,
                CASE
                    WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                    WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                    ELSE ''
                END AS glosa,
                CASE
                    WHEN m.tipo_movimiento = 'Compras' THEN 8
                    WHEN m.tipo_movimiento = 'Ventas' THEN 14
                    ELSE NULL
                END AS codigo_del_libro,
                DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo_documento,
                ac.numero_documento AS numero_documento_sustentatorio,
                ac.codigo_cuenta,
                ac.denominacion,
                ac.debe,
                ac.haber
            FROM asientos_contables ac
            JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
            WHERE ac.fecha::date = %s::date
            ORDER BY fecha, numero_correlativo, ac.id;
        """, (fecha,))
        
        movimientos = cursor.fetchall()

        for movimiento in movimientos:
            total_debe += movimiento['debe'] or 0
            total_haber += movimiento['haber'] or 0

    conexion.close()
    return movimientos, total_debe, total_haber

def obtener_libro_diario_por_fecha(fecha_inicio, fecha_fin=None):
    conexion = obtener_conexion()
    movimientos_agrupados = []
    total_debe = 0
    total_haber = 0
    try:
        with conexion.cursor(cursor_factory=DictCursor) as cursor:
            if fecha_fin:
                query = """
                    SELECT
                        DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                        TO_CHAR(ac.fecha, 'DD/MM/YYYY') AS fecha,
                        CASE
                            WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                            WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                            ELSE ''
                        END AS glosa,
                        CASE
                            WHEN m.tipo_movimiento = 'Compras' THEN 8
                            WHEN m.tipo_movimiento = 'Ventas' THEN 14
                            ELSE NULL
                        END AS codigo_del_libro,
                        DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo_documento,
                        ac.numero_documento AS numero_documento_sustentatorio,
                        ac.codigo_cuenta,
                        ac.denominacion,
                        ac.debe,
                        ac.haber
                    FROM asientos_contables ac
                    JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
                    WHERE ac.fecha::date BETWEEN %s AND %s
                    ORDER BY ac.fecha ASC, numero_correlativo, ac.id;
                """
                cursor.execute(query, (fecha_inicio, fecha_fin))
            else:
                query = """
                    SELECT
                        DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                        TO_CHAR(ac.fecha, 'DD/MM/YYYY') AS fecha,
                        CASE
                            WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                            WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                            ELSE ''
                        END AS glosa,
                        CASE
                            WHEN m.tipo_movimiento = 'Compras' THEN 8
                            WHEN m.tipo_movimiento = 'Ventas' THEN 14
                            ELSE NULL
                        END AS codigo_del_libro,
                        DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo_documento,
                        ac.numero_documento AS numero_documento_sustentatorio,
                        ac.codigo_cuenta,
                        ac.denominacion,
                        ac.debe,
                        ac.haber
                    FROM asientos_contables ac
                    JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
                    WHERE ac.fecha::date = %s
                    ORDER BY ac.fecha ASC, numero_correlativo, ac.id;
                """
                cursor.execute(query, (fecha_inicio,))
            movimientos = cursor.fetchall()
            agrupado = {}
            for movimiento in movimientos:
                numero_correlativo = movimiento["numero_correlativo"]
                if numero_correlativo not in agrupado:
                    agrupado[numero_correlativo] = {
                        "numero_correlativo": numero_correlativo,
                        "fecha": movimiento["fecha"],
                        "glosa": movimiento["glosa"],
                        "codigo_del_libro": movimiento["codigo_del_libro"],
                        "numero_correlativo_documento": movimiento["numero_correlativo_documento"],
                        "numero_documento_sustentatorio": movimiento["numero_documento_sustentatorio"],
                        "cuentas": []
                    }
                agrupado[numero_correlativo]["cuentas"].append({
                    "codigo_cuenta": movimiento["codigo_cuenta"],
                    "denominacion": movimiento["denominacion"],
                    "debe": movimiento["debe"],
                    "haber": movimiento["haber"]
                })
                total_debe += movimiento["debe"] or 0
                total_haber += movimiento["haber"] or 0
            movimientos_agrupados = list(agrupado.values())
    except Exception as e:
        print(f"Error al obtener los movimientos del libro diario: {e}")
    finally:
        conexion.close()
    return movimientos_agrupados, total_debe, total_haber

def obtener_movimientos_libro_diario(fecha):
    conexion = obtener_conexion()
    movimientos = []
    total_debe = 0
    total_haber = 0
    try:
        with conexion.cursor(cursor_factory=DictCursor) as cursor:
            cursor.execute("""
                SELECT
                    DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                    TO_CHAR(ac.fecha, 'DD/MM/YYYY') AS fecha,
                    CASE
                        WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                        WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                        ELSE ''
                    END AS glosa,
                    CASE
                        WHEN m.tipo_movimiento = 'Compras' THEN 8
                        WHEN m.tipo_movimiento = 'Ventas' THEN 14
                        ELSE NULL
                    END AS codigo_del_libro,
                    ac.numero_documento AS numero_documento_sustentatorio,
                    ac.codigo_cuenta,
                    ac.denominacion,
                    ac.debe,
                    ac.haber
                FROM asientos_contables ac
                JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
                WHERE ac.fecha::date = %s::date
                ORDER BY fecha, numero_correlativo, ac.id;
            """, (fecha,))
            movimientos = cursor.fetchall()
            for movimiento in movimientos:
                total_debe += movimiento['debe'] or 0
                total_haber += movimiento['haber'] or 0
    except Exception as e:
        print(f"Error al obtener los movimientos del libro diario: {e}")
    finally:
        conexion.close()
    return movimientos, total_debe, total_haber


def obtener_libro_caja(mes, año):
    try:
        mes = int(mes)
        año = int(año)
    except ValueError:
        print("Error: mes o año no son enteros válidos")
        return [], 0, 0
    conexion = obtener_conexion()
    movimientos_agrupados = []
    total_deudor = 0
    total_acreedor = 0
    with conexion.cursor(cursor_factory=DictCursor) as cursor:
        try:
            query = sql.SQL("""
                SELECT
                    DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                    TO_CHAR(ac.fecha, 'DD/MM/YYYY') AS fecha_operacion,
                    CASE
                        WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                        WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                        ELSE 'Descripción no especificada'
                    END AS descripcion_operacion,
                    ac.codigo_cuenta AS codigo_cuenta_asociada,
                    ac.denominacion AS denominacion_cuenta_asociada,
                    COALESCE(ac.debe, 0) AS saldo_deudor,
                    COALESCE(ac.haber, 0) AS saldo_acreedor
                FROM asientos_contables ac
                JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
                WHERE 
                    (ac.codigo_cuenta LIKE '10%')
                    AND EXTRACT(MONTH FROM ac.fecha) = {mes}
                    AND EXTRACT(YEAR FROM ac.fecha) = {año}
                ORDER BY fecha_operacion, numero_correlativo, ac.id;
            """).format(mes=sql.Literal(mes), año=sql.Literal(año))
            cursor.execute(query)
            movimientos = cursor.fetchall()
            agrupado = {}
            for movimiento in movimientos:
                numero_correlativo = movimiento["numero_correlativo"]
                if numero_correlativo not in agrupado:
                    agrupado[numero_correlativo] = {
                        "numero_correlativo": numero_correlativo,
                        "fecha_operacion": movimiento["fecha_operacion"],
                        "descripcion_operacion": movimiento["descripcion_operacion"],
                        "cuentas": []
                    }
                agrupado[numero_correlativo]["cuentas"].append({
                    "codigo_cuenta_asociada": movimiento["codigo_cuenta_asociada"],
                    "denominacion_cuenta_asociada": movimiento["denominacion_cuenta_asociada"],
                    "saldo_deudor": movimiento["saldo_deudor"],
                    "saldo_acreedor": movimiento["saldo_acreedor"]
                })
                total_deudor += movimiento["saldo_deudor"]
                total_acreedor += movimiento["saldo_acreedor"]
            movimientos_agrupados = list(agrupado.values())
        except Exception as e:
            print("Error en la consulta SQL con `psycopg2.sql`:", e)
            movimientos_agrupados, total_deudor, total_acreedor = [], 0, 0
    conexion.close()
    return movimientos_agrupados, total_deudor, total_acreedor


def obtener_cuentas_distintas():
    conexion = obtener_conexion()
    cuentas = []

    with conexion.cursor(cursor_factory=DictCursor) as cursor:
        cursor.execute("""
            SELECT DISTINCT ac.codigo_cuenta, ac.denominacion
            FROM asientos_contables ac
            JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
            WHERE EXTRACT(MONTH FROM ac.fecha) = 11
              AND EXTRACT(YEAR FROM ac.fecha) = 2024
              AND (ac.debe IS NOT NULL AND ac.debe != 0 OR ac.haber IS NOT NULL AND ac.haber != 0)
            ORDER BY ac.codigo_cuenta;
        """)
        
        cuentas = cursor.fetchall()

    conexion.close()
    return cuentas

def obtener_libro_mayor(mes, año, cuenta):
    conexion = obtener_conexion()
    movimientos = []

    with conexion.cursor(cursor_factory=DictCursor) as cursor:
        cursor.execute("""
            SELECT fecha::date, numero_correlativo, glosa, debe as deudor, haber as acreedor
            FROM (
                SELECT
                    DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                    ac.fecha,
                    CASE
                        WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                        WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                        ELSE ''
                    END AS glosa,
                    ac.debe,
                    ac.haber
                FROM asientos_contables ac
                JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
                WHERE EXTRACT(MONTH FROM ac.fecha) = %s
                AND EXTRACT(YEAR FROM ac.fecha) = %s
                AND ac.codigo_cuenta = %s
                AND (ac.debe IS NOT NULL AND ac.debe != 0 OR ac.haber IS NOT NULL AND ac.haber != 0)
                ORDER BY fecha, numero_correlativo, ac.id
            ) AS subquery;
        """, (mes, año, cuenta))
        
        movimientos = cursor.fetchall()
        print("Movimientos encontrados:", movimientos)  # Mensaje de depuración para ver el resultado de la consulta.

    conexion.close()
    return movimientos

def generar_libro_mayor_excel(mes, año, cuenta):
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor()

        consulta_cuenta = """
            SELECT codigo, descripcion
            FROM public.cuentas
            WHERE codigo = %s
        """
        cursor.execute(consulta_cuenta, (cuenta,))
        cuenta_resultado = cursor.fetchone()

        if cuenta_resultado:
            cuenta_codigo, cuenta_descripcion = cuenta_resultado
            cuenta_formateada = f"{cuenta_codigo} - {cuenta_descripcion}"
        else:
            cuenta_formateada = "Cuenta no encontrada"

        consulta = """
            SELECT fecha, numero_correlativo, glosa, debe as deudor, haber as acreedor
            FROM (
                SELECT
                    DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                    ac.fecha,
                    CASE
                        WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                        WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                        ELSE ''
                    END AS glosa,
                    ac.debe,
                    ac.haber
                FROM asientos_contables ac
                JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
                WHERE EXTRACT(MONTH FROM ac.fecha) = %s
                AND EXTRACT(YEAR FROM ac.fecha) = %s
                AND ac.codigo_cuenta = %s
                AND (ac.debe IS NOT NULL AND ac.debe != 0 OR ac.haber IS NOT NULL AND ac.haber != 0)
                ORDER BY fecha, numero_correlativo, ac.id
            ) AS subquery;
        """
        cursor.execute(consulta, (mes, año, cuenta))
        resultados = cursor.fetchall()

        ruta_plantilla = 'plantillas/LibroMayor.xlsx'
        workbook = load_workbook(ruta_plantilla)
        hoja = workbook.active

        periodo = f"{año}-{mes.zfill(2)}"
        hoja.cell(row=3, column=2, value=periodo)

        hoja.cell(row=6, column=3, value=cuenta_formateada)

        borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fuente_estandar = Font(name='Arial', size=10)

        fila_base = 11
        alto_fila_base = hoja.row_dimensions[fila_base].height
        fila_inicial = fila_base
        total_deudor = 0
        total_acreedor = 0

        columnas_con_borde = list(range(1, 6))

        for fila, registro in enumerate(resultados, start=fila_inicial):
            hoja.row_dimensions[fila].height = alto_fila_base
            celdas = [
                (1, registro[0]),  # Fecha
                (2, registro[1]),  # Número Correlativo
                (3, registro[2]),  # Glosa
                (4, registro[3]),  # Deudor
                (5, registro[4])   # Acreedor
            ]

            for col in columnas_con_borde:
                celda = hoja.cell(row=fila, column=col)
                celda.border = borde
                celda.font = fuente_estandar

            for col, valor in celdas:
                celda = hoja.cell(row=fila, column=col, value=valor)
                celda.font = fuente_estandar

                if col in (4, 5):  # Columnas Deudor y Acreedor
                    celda.alignment = Alignment(horizontal='right', vertical='center')
                    celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                elif col == 1:  # Fecha
                    celda.number_format = FORMAT_DATE_DDMMYY
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                elif col == 2:
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    celda.alignment = Alignment(horizontal='left', vertical='center')

                if col == 4:
                    total_deudor += valor or 0
                elif col == 5:
                    total_acreedor += valor or 0

        fila_totales = fila_inicial + len(resultados)
        hoja.row_dimensions[fila_totales].height = alto_fila_base
        hoja.cell(row=fila_totales, column=3, value="TOTALES").border = borde
        hoja.cell(row=fila_totales, column=3).alignment = Alignment(horizontal='center', vertical='center')
        hoja.cell(row=fila_totales, column=3).font = Font(name='Arial', size=10, bold=True)

        hoja.cell(row=fila_totales, column=4, value=total_deudor).border = borde
        hoja.cell(row=fila_totales, column=4).number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        hoja.cell(row=fila_totales, column=4).alignment = Alignment(horizontal='right', vertical='center')
        hoja.cell(row=fila_totales, column=4).font = fuente_estandar

        hoja.cell(row=fila_totales, column=5, value=total_acreedor).border = borde
        hoja.cell(row=fila_totales, column=5).number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        hoja.cell(row=fila_totales, column=5).alignment = Alignment(horizontal='right', vertical='center')
        hoja.cell(row=fila_totales, column=5).font = fuente_estandar

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        nombre_archivo = f'libro_mayor_{año}_{mes}_{cuenta}.xlsx'
        return send_file(
            output,
            download_name=nombre_archivo,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error al generar el libro mayor: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close()
            conexion.close()

def generar_libro_mayor_pdf(mes, año, cuenta):
    try:
        # Establecer la conexión a la base de datos
        conexion = obtener_conexion()
        cursor = conexion.cursor(cursor_factory=DictCursor)

        # Consulta SQL para obtener los movimientos del Libro Mayor
        consulta = """
            SELECT 
                TO_CHAR(fecha, 'DD/MM/YYYY') AS fecha, 
                numero_correlativo, 
                glosa, 
                denominacion,
                debe AS deudor, 
                haber AS acreedor
            FROM (
                SELECT
                    DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                    ac.fecha,
                    CASE
                        WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                        WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                        ELSE 'Operación no especificada'
                    END AS glosa,
                    ac.denominacion,
                    ac.debe,
                    ac.haber
                FROM asientos_contables ac
                JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
                WHERE EXTRACT(MONTH FROM ac.fecha) = %s
                AND EXTRACT(YEAR FROM ac.fecha) = %s
                AND ac.codigo_cuenta = %s
                AND (ac.debe IS NOT NULL AND ac.debe != 0 OR ac.haber IS NOT NULL AND ac.haber != 0)
                ORDER BY fecha, numero_correlativo, ac.id
            ) AS subquery;
        """
        cursor.execute(consulta, (mes, año, cuenta))
        movimientos = cursor.fetchall()

        if not movimientos:
            return jsonify({'error': 'No se encontraron datos para el período y cuenta seleccionados.'}), 404

        # Procesar los datos
        procesados = []
        total_debe = 0
        total_haber = 0

        for movimiento in movimientos:
            procesados.append(
                (
                    movimiento["fecha"],  # Fecha ya formateada como DD/MM/YYYY
                    movimiento["numero_correlativo"],
                    movimiento["glosa"],
                    f"{movimiento['deudor']:,.2f}" if movimiento["deudor"] else "-",
                    f"{movimiento['acreedor']:,.2f}" if movimiento["acreedor"] else "-"
                )
            )
            total_debe += movimiento["deudor"] or 0
            total_haber += movimiento["acreedor"] or 0

        # Crear el PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=10,
            leftMargin=10,
            topMargin=20,
            bottomMargin=20
        )

        styles = getSampleStyleSheet()
        style_title = styles["Title"]
        
        # Título del documento
        titulo = Paragraph(f"<b>Libro Mayor - Detalle de las operaciones</b>", style_title)

        # Cabecera con el Periodo, RUC, Razón Social en negrita
        ruc = "20612188930"  # RUC fijo
        razon_social = "DECO ELERA S.A.C."  # Razón Social
        fecha_texto = f"{mes}/{año}"

        # Concatenación de cuenta y denominación
        cuenta_denominacion = f"<b>Cuenta: {cuenta} - {movimientos[0]['denominacion']}</b>"

        # Usar <b> para negrita en los textos de la cabecera
        cabecera = [
            f"<b>Periodo: {fecha_texto}</b>",
            f"<b>RUC: {ruc}</b>",
            f"<b>Razón Social: {razon_social}</b>",
            cuenta_denominacion  # Aquí concatenamos cuenta y denominación
        ]
        
        # Añadir título y cabecera a los elementos
        elementos = [titulo, Spacer(1, 12)]  # Espacio después del título

        for linea in cabecera:
            elementos.append(Paragraph(linea, styles['Normal']))
        
        elementos.append(Spacer(1, 12))  # Espacio después de la cabecera

        # Construir la tabla
        encabezados = ["Fecha", "N° Correlativo", "Descripción Operación", "Deudor", "Acreedor"]
        tabla_datos = [encabezados]

        for registro in procesados:
            tabla_datos.append(list(registro))

        # Agregar totales
        tabla_datos.append([ 
            "", "", "Totales",
            f"{total_debe:,.2f}" if total_debe else "-",
            f"{total_haber:,.2f}" if total_haber else "-"
        ])

        # Configurar tabla
        tabla = Table(
            tabla_datos,
            colWidths=[80, 100, 300, 80, 80]
        )
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -2), 'LEFT'),  # Descripción alineada a la izquierda
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),  # Montos alineados a la derecha
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))

        # Añadir tabla a los elementos
        elementos.append(tabla)

        # Generar el PDF con la marca de agua en cada página
        doc.build(elementos, onFirstPage=add_watermark_image, onLaterPages=add_watermark_image)

        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"Libro_Mayor_{mes}_{año}_{cuenta}.pdf",
            mimetype="application/pdf"
        )

    except Exception as e:
        print(f"Error al generar el PDF: {e}")
        return jsonify({"error": str(e)}), 500

    finally:
        if conexion:
            cursor.close()
            conexion.close()




load_dotenv()

def generar_libro_caja_excel(mes, anio):
    print(f"mes: '{mes}', año: '{anio}'")
    try:
        db_url = f"postgresql://{os.getenv('POSTGRES_USER')}:{os.getenv('POSTGRES_PASSWORD')}@{os.getenv('POSTGRES_HOST')}:{os.getenv('POSTGRES_PORT')}/{os.getenv('POSTGRES_DB')}"
        engine = create_engine(db_url)
        with engine.connect() as conexion:
            consulta = """
            SELECT 
                DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                TO_CHAR(ac.fecha, 'DD/MM/YYYY') AS fecha_operacion,
                CASE
                    WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                    WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                    ELSE 'Descripción no especificada'
                END AS descripcion_operacion,
                ac.codigo_cuenta AS codigo_cuenta_asociada,
                ac.denominacion AS denominacion_cuenta_asociada,
                COALESCE(ac.debe, 0) AS saldo_deudor,
                COALESCE(ac.haber, 0) AS saldo_acreedor
            FROM asientos_contables ac
            JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
            WHERE 
                (ac.codigo_cuenta LIKE '10%')
                AND EXTRACT(MONTH FROM ac.fecha) = :mes
                AND EXTRACT(YEAR FROM ac.fecha) = :anio
            ORDER BY ac.fecha, numero_correlativo, codigo_cuenta_asociada;
            """
            # Ejecuta la consulta y obtiene los resultados
            resultados = conexion.execute(text(consulta), {"mes": mes, "anio": anio}).fetchall()

            if not resultados:
                return jsonify({'error': 'No se encontraron resultados.'}), 404

            # Cargar la plantilla de Excel
            ruta_plantilla = 'plantillas/LibroCaja.xlsx'
            workbook = load_workbook(ruta_plantilla)
            hoja = workbook.active
            
            periodo = f"{anio}-{str(mes).zfill(2)}"
            hoja.cell(row=3, column=2, value=periodo)

            # Definir estilos
            borde = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            fuente_estandar = Font(name='Arial', size=10)

            fila_base = 9
            alto_fila_base = hoja.row_dimensions[fila_base].height  # Alto de fila base
            fila_inicial = fila_base
            total_deudor = 0
            total_acreedor = 0
            columnas_con_borde = list(range(1, 8))  # Ajusta según las columnas que usas en la plantilla

            # Establecer el alto de todas las filas
            for fila in range(fila_inicial, fila_inicial + len(resultados) + 1):
                hoja.row_dimensions[fila].height = alto_fila_base

            for fila, registro in enumerate(resultados, start=fila_inicial):
                if len(registro) < 7:
                    print(f"Error: El registro tiene menos de 7 columnas: {registro}")
                    continue  # Salta el registro incorrecto

                celdas = [
                    (1, registro[0]),  # Correlativo
                    (2, registro[1]),  # Fecha operación
                    (3, registro[2]),  # Descripción operación
                    (4, registro[3]),  # Código cuenta
                    (5, registro[4]),  # Denominación cuenta
                    (6, registro[5]),  # Saldo deudor
                    (7, registro[6])   # Saldo acreedor
                ]

                # Aplica borde y fuente estándar a todas las celdas de la fila
                for col in columnas_con_borde:
                    celda = hoja.cell(row=fila, column=col)
                    celda.border = borde
                    celda.font = fuente_estandar

                # Asigna los valores a las celdas correspondientes
                for col, valor in celdas:
                    celda = hoja.cell(row=fila, column=col, value=valor)
                    celda.font = fuente_estandar

                    # Formatear las celdas de acuerdo a su tipo
                    if col in (6, 7):  # Saldo deudor y saldo acreedor
                        celda.alignment = Alignment(horizontal='right', vertical='center')
                        celda.number_format = '###,##0.00'
                    elif col == 2:  # Fecha de operación
                        celda.number_format = 'DD/MM/YYYY'
                        celda.alignment = Alignment(horizontal='center', vertical='center')
                    elif col == 1:
                        celda.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        celda.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    # Sumar los totales deudor y acreedor
                    if col == 6:
                        total_deudor += valor
                    elif col == 7:
                        total_acreedor += valor

            # Escribir los totales en la fila de "Totales"
            fila_totales = fila_inicial + len(resultados)
            hoja.row_dimensions[fila_totales].height = alto_fila_base  # Mantener el mismo alto para la fila de totales
            celda_totales = hoja.cell(row=fila_totales, column=5, value="TOTALES")
            celda_totales.border = borde
            celda_totales.alignment = Alignment(horizontal='center', vertical='center')
            celda_totales.font = Font(name='Arial', size=10, bold=True)

            total_deudor_celda = hoja.cell(row=fila_totales, column=6, value=total_deudor)
            total_deudor_celda.border = borde
            total_deudor_celda.number_format = '###,##0.00'
            total_deudor_celda.alignment = Alignment(horizontal='right', vertical='center')
            total_deudor_celda.font = fuente_estandar

            total_acreedor_celda = hoja.cell(row=fila_totales, column=7, value=total_acreedor)
            total_acreedor_celda.border = borde
            total_acreedor_celda.number_format = '###,##0.00'
            total_acreedor_celda.alignment = Alignment(horizontal='right', vertical='center')
            total_acreedor_celda.font = fuente_estandar

            # Aplicar bordes a las celdas de totales
            for col in columnas_con_borde:
                hoja.cell(row=fila_totales, column=col).border = borde

            # Guardar el archivo Excel en memoria
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            nombre_archivo = f'libro_caja_{anio}_{mes}.xlsx'

            # Enviar el archivo como respuesta
            return send_file(
                output,
                download_name=nombre_archivo,
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    except Exception as e:
        print(f"Error al generar el libro de caja: {e}")
        return jsonify({'error': str(e)}), 500

def generar_excel_todas_las_cuentas(mes, año):
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor()

        cuentas = obtener_cuentas_distintas()  # Lista de cuentas únicas
        ruta_plantilla = 'plantillas/LibroMayor.xlsx'
        workbook = load_workbook(ruta_plantilla)

        borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fuente_estandar = Font(name='Arial', size=10)
        alto_fila_base = workbook.active.row_dimensions[11].height  # Usamos la altura de la fila base de la plantilla
        columnas_con_borde = list(range(1, 6))

        for cuenta in cuentas:
            codigo_cuenta = cuenta['codigo_cuenta']

            consulta_cuenta = """
                SELECT codigo, descripcion
                FROM public.cuentas
                WHERE codigo = %s
            """
            cursor.execute(consulta_cuenta, (codigo_cuenta,))
            cuenta_resultado = cursor.fetchone()

            if cuenta_resultado:
                cuenta_codigo, cuenta_descripcion = cuenta_resultado
                cuenta_formateada = f"{cuenta_codigo} - {cuenta_descripcion}"
            else:
                cuenta_formateada = "Cuenta no encontrada"

            hoja = workbook.copy_worksheet(workbook.active)
            hoja.title = f"Cuenta {codigo_cuenta}"

            periodo = f"{año}-{mes.zfill(2)}"
            hoja.cell(row=3, column=2, value=periodo)
            hoja.cell(row=6, column=3, value=cuenta_formateada)

            consulta = """
                SELECT fecha, numero_correlativo, glosa, debe as deudor, haber as acreedor
                FROM (
                    SELECT
                        DENSE_RANK() OVER (ORDER BY ac.numero_asiento) AS numero_correlativo,
                        ac.fecha,
                        CASE
                            WHEN m.tipo_movimiento = 'Ventas' THEN 'Por la venta de mercadería'
                            WHEN m.tipo_movimiento = 'Compras' THEN 'Por la compra de insumos'
                            ELSE ''
                        END AS glosa,
                        ac.debe,
                        ac.haber
                    FROM asientos_contables ac
                    JOIN movimientos m ON ac.numero_asiento = m.movimiento_id
                    WHERE EXTRACT(MONTH FROM ac.fecha) = %s
                    AND EXTRACT(YEAR FROM ac.fecha) = %s
                    AND ac.codigo_cuenta = %s
                    AND (ac.debe IS NOT NULL AND ac.debe != 0 OR ac.haber IS NOT NULL AND ac.haber != 0)
                    ORDER BY fecha, numero_correlativo, ac.id
                ) AS subquery;
            """
            cursor.execute(consulta, (mes, año, codigo_cuenta))
            resultados = cursor.fetchall()

            fila_inicial = 11
            total_deudor = 0
            total_acreedor = 0

            for fila, registro in enumerate(resultados, start=fila_inicial):
                hoja.row_dimensions[fila].height = alto_fila_base
                celdas = [
                    (1, registro[0]),  # Fecha
                    (2, registro[1]),  # Número Correlativo
                    (3, registro[2]),  # Glosa
                    (4, registro[3]),  # Deudor
                    (5, registro[4])   # Acreedor
                ]

                for col, valor in celdas:
                    celda = hoja.cell(row=fila, column=col, value=valor)
                    celda.border = borde
                    celda.font = fuente_estandar

                    if col in (4, 5):  # Columnas Deudor y Acreedor
                        celda.alignment = Alignment(horizontal='right', vertical='center')
                        celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                    elif col == 1:  # Fecha
                        celda.number_format = FORMAT_DATE_DDMMYY
                        celda.alignment = Alignment(horizontal='center', vertical='center')
                    elif col == 2:
                        celda.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        celda.alignment = Alignment(horizontal='left', vertical='center')

                    if col == 4:
                        total_deudor += valor or 0
                    elif col == 5:
                        total_acreedor += valor or 0

            fila_totales = fila_inicial + len(resultados)
            hoja.row_dimensions[fila_totales].height = alto_fila_base
            hoja.cell(row=fila_totales, column=3, value="TOTALES").border = borde
            hoja.cell(row=fila_totales, column=3).alignment = Alignment(horizontal='center', vertical='center')
            hoja.cell(row=fila_totales, column=3).font = Font(name='Arial', size=10, bold=True)

            hoja.cell(row=fila_totales, column=4, value=total_deudor).border = borde
            hoja.cell(row=fila_totales, column=4).number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            hoja.cell(row=fila_totales, column=4).alignment = Alignment(horizontal='right', vertical='center')
            hoja.cell(row=fila_totales, column=4).font = fuente_estandar

            hoja.cell(row=fila_totales, column=5, value=total_acreedor).border = borde
            hoja.cell(row=fila_totales, column=5).number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            hoja.cell(row=fila_totales, column=5).alignment = Alignment(horizontal='right', vertical='center')
            hoja.cell(row=fila_totales, column=5).font = fuente_estandar

        workbook.remove(workbook.active)  # Eliminar la hoja base de la plantilla
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        conexion.close()
        return output

    except Exception as e:
        print(f"Error al generar el libro mayor para todas las cuentas: {e}")
        return None