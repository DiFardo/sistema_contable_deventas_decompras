import psycopg2
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY, FORMAT_NUMBER_COMMA_SEPARATED1
from bd_conexion import obtener_conexion

def generar_registro_venta_excel(mes, anio, ruta_plantilla, ruta_salida):
    try:
        # Establecer conexión con la base de datos
        conexion = obtener_conexion()
        cursor = conexion.cursor()

        # Ejecutar la consulta SQL
        consulta = """
        SELECT 
            ROW_NUMBER() OVER(ORDER BY v.serie_comprobante, v.numero_comprobante) AS correlativo,
            TO_CHAR(v.fecha, 'DD/MM/YYYY') AS fecha_emision,
            CASE WHEN v.tipo_comprobante = 'Boleta' THEN '03' ELSE '01' END AS tipo_comprobante,
            v.serie_comprobante,
            v.numero_comprobante,
            CASE 
                WHEN v.tipo_documento = 'DNI' THEN '1' 
                WHEN v.tipo_documento = 'Carnet de extranjería' THEN '4'
                WHEN v.tipo_documento = 'RUC' THEN '6'
                WHEN v.tipo_documento = 'Pasaporte' THEN '7'
                ELSE ''
            END AS tipo_documento,
            v.numero_documento,
            v.usuario,
            SUM(v.sub_sin_igv) AS base_imponible,
            SUM(v.igv) AS igv,
            SUM(v.subtotal) AS total_comprobante
        FROM ventas_contables v
        WHERE EXTRACT(MONTH FROM v.fecha) = %s AND EXTRACT(YEAR FROM v.fecha) = %s
        GROUP BY v.serie_comprobante, v.numero_comprobante, v.tipo_documento, 
                 v.numero_documento, v.usuario, v.tipo_comprobante, v.fecha
        ORDER BY v.serie_comprobante, v.numero_comprobante;
        """
        cursor.execute(consulta, (mes, anio))

        # Obtener los resultados de la consulta
        resultados = cursor.fetchall()

        # Cargar la plantilla de Excel
        workbook = load_workbook(ruta_plantilla)
        hoja = workbook.active

        # Estilo de bordes para las celdas
        borde = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Copiar el estilo de la fila base (por ejemplo, la fila 12)
        fila_base = 12
        alto_fila_base = hoja.row_dimensions[fila_base].height

        # Insertar los datos en la plantilla
        fila_inicial = fila_base  # Inicio de inserción de datos
        total_base_imponible = 0
        total_igv = 0
        total_comprobante = 0

        for fila, registro in enumerate(resultados, start=fila_inicial):
            hoja.row_dimensions[fila].height = alto_fila_base  # Mantener el mismo alto de fila

            # Insertar valores y aplicar estilos a las celdas
            celdas = [
                (1, registro[0]),  # correlativo
                (2, registro[1]),  # fecha_emision
                (4, registro[2]),  # tipo_comprobante
                (5, registro[3]),  # serie_comprobante
                (6, registro[4]),  # numero_comprobante
                (7, registro[5]),  # tipo_documento
                (8, registro[6]),  # numero_documento
                (9, registro[7]),  # usuario
                (11, registro[8]),  # base_imponible
                (15, registro[9]),  # igv
                (17, registro[10])  # total_comprobante
            ]

            for col, valor in celdas:
                celda = hoja.cell(row=fila, column=col, value=valor)
                celda.border = borde
                celda.alignment = Alignment(horizontal='center', vertical='center')

                # Aplicar formatos específicos
                if col == 2:  # Columna de fecha
                    celda.number_format = FORMAT_DATE_DDMMYY
                elif col in (11, 15, 17):  # Columnas de números
                    celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1

            # Sumar los totales
            total_base_imponible += registro[8]
            total_igv += registro[9]
            total_comprobante += registro[10]

        # Escribir los totales en la fila de "Totales"
        fila_totales = fila_inicial + len(resultados)
        hoja.row_dimensions[fila_totales].height = alto_fila_base
        celda_totales = hoja.cell(row=fila_totales, column=9, value="TOTALES")
        celda_totales.border = borde
        celda_totales.alignment = Alignment(horizontal='center', vertical='center')
        celda_totales.font = Font(bold=True)  # Aplicar negrita

        # Totales en columnas de base imponible, IGV y total comprobante
        total_base_imponible_celda = hoja.cell(row=fila_totales, column=11, value=total_base_imponible)
        total_base_imponible_celda.border = borde
        total_base_imponible_celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1

        total_igv_celda = hoja.cell(row=fila_totales, column=15, value=total_igv)
        total_igv_celda.border = borde
        total_igv_celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1

        total_comprobante_celda = hoja.cell(row=fila_totales, column=17, value=total_comprobante)
        total_comprobante_celda.border = borde
        total_comprobante_celda.number_format = FORMAT_NUMBER_COMMA_SEPARATED1

        # Guardar el nuevo archivo de Excel
        workbook.save(ruta_salida)
        print(f"Registro de ventas generado exitosamente en: {ruta_salida}")

    except Exception as e:
        print(f"Error al generar el registro de ventas: {e}")
    finally:
        # Cerrar la conexión con la base de datos
        if conexion:
            cursor.close()
            conexion.close()

# Ejemplo de uso
mes = 10  # Octubre
anio = 2024  # Año 2024
ruta_plantilla = 'plantillas/RegistroVentas.xlsx'
ruta_salida = f'registro_ventas_{anio}_{mes}.xlsx'

generar_registro_venta_excel(mes, anio, ruta_plantilla, ruta_salida)